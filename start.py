# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, messagebox
import json
import os
import re
import time
import threading

FILE_LOCATION = os.path.dirname(os.path.realpath(__file__))
DEPS = os.path.join(FILE_LOCATION, "Dependencies")
KLIC = os.path.join(FILE_LOCATION, "KLIC")


def load_data():
    with open(os.path.join(DEPS, "Opdrachtgevers.json"), "r", encoding="utf-8") as f:
        return json.load(f)


def set_cell(sheet, cell_ref, value):
    """Schrijf naar een cel, ook als die deel uitmaakt van een samengevoegde reeks."""
    from openpyxl.utils import column_index_from_string
    col_str = re.match(r'([A-Z]+)', cell_ref.upper()).group(1)
    row = int(re.match(r'[A-Z]+(\d+)', cell_ref.upper()).group(1))
    col = column_index_from_string(col_str)
    for merged in sheet.merged_cells.ranges:
        if (merged.min_row <= row <= merged.max_row and
                merged.min_col <= col <= merged.max_col):
            # Schrijf naar de top-left cel van de merge
            sheet.cell(row=merged.min_row, column=merged.min_col).value = value
            return
    sheet[cell_ref] = value


def add_image_to_range(sheet, img_path, from_cell, to_cell):
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    from openpyxl.utils import column_index_from_string
    from openpyxl.drawing.image import Image

    def parse(ref):
        m = re.match(r'([A-Z]+)(\d+)', ref.upper())
        return column_index_from_string(m.group(1)) - 1, int(m.group(2)) - 1

    fc, fr = parse(from_cell)
    tc, tr = parse(to_cell)
    anchor = TwoCellAnchor()
    anchor._from = AnchorMarker(col=fc, colOff=0, row=fr, rowOff=0)
    anchor.to    = AnchorMarker(col=tc, colOff=0, row=tr, rowOff=0)
    anchor.editAs = 'oneCell'
    img = Image(img_path)
    img.anchor = anchor
    sheet.add_image(img)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Productblad Generator — B-Advice")
        self.resizable(False, False)
        self.data = load_data()
        self.web_driver = None
        self.google_url = ""
        self.location_code = None
        self.output_dir = None
        self._build()
        self._fill_opdrachtgevers()

    # ------------------------------------------------------------------ UI --
    def _build(self):
        pad = dict(padx=6, pady=4)

        # --- rij 1: opdrachtgever / plaats / wijk ---
        f1 = ttk.LabelFrame(self, text="Locatie")
        f1.grid(row=0, column=0, columnspan=2, sticky="ew", **pad)

        ttk.Label(f1, text="Opdrachtgever:").grid(row=0, column=0, **pad)
        ttk.Label(f1, text="Plaats:").grid(row=0, column=1, **pad)
        ttk.Label(f1, text="Wijk:").grid(row=0, column=2, **pad)

        self.cb_opdrachtgever = ttk.Combobox(f1, state="readonly", width=22)
        self.cb_opdrachtgever.grid(row=1, column=0, **pad)
        self.cb_opdrachtgever.bind("<<ComboboxSelected>>", lambda _: self._update_plaats())

        self.cb_plaats = ttk.Combobox(f1, state="readonly", width=22)
        self.cb_plaats.grid(row=1, column=1, **pad)
        self.cb_plaats.bind("<<ComboboxSelected>>", lambda _: self._update_wijk())

        self.cb_wijk = ttk.Combobox(f1, state="readonly", width=22)
        self.cb_wijk.grid(row=1, column=2, **pad)

        # --- rij 2: adres ---
        f2 = ttk.LabelFrame(self, text="Adres")
        f2.grid(row=1, column=0, columnspan=2, sticky="ew", **pad)

        labels2 = ["Straat:", "Huisnummer:", "Toevoeging:", "Postcode:", "Coordinaten:"]
        for i, lbl in enumerate(labels2):
            ttk.Label(f2, text=lbl).grid(row=0, column=i, **pad)

        self.e_straat      = ttk.Entry(f2, width=18); self.e_straat.grid(row=1, column=0, **pad)
        self.e_huisnummer  = ttk.Entry(f2, width=8);  self.e_huisnummer.grid(row=1, column=1, **pad)
        self.e_toevoeging  = ttk.Entry(f2, width=6);  self.e_toevoeging.grid(row=1, column=2, **pad)
        self.e_postcode    = ttk.Entry(f2, width=10); self.e_postcode.grid(row=1, column=3, **pad)

        # Coordinaten veld + knop om automatisch op te halen
        coord_frame = ttk.Frame(f2)
        coord_frame.grid(row=1, column=4, **pad)
        self.e_coordinaten = ttk.Entry(coord_frame, width=26)
        self.e_coordinaten.pack(side="left")
        ttk.Button(coord_frame, text="📍", width=3,
                   command=self._haal_coordinaten).pack(side="left", padx=(2, 0))

        # Auto-ophalen als postcode + huisnummer zijn ingevuld
        self.e_postcode.bind("<FocusOut>", lambda _: self._auto_coordinaten())
        self.e_huisnummer.bind("<FocusOut>", lambda _: self._auto_coordinaten())

        # --- rij 3: extra info ---
        f3 = ttk.LabelFrame(self, text="Extra")
        f3.grid(row=2, column=0, columnspan=2, sticky="ew", **pad)

        ttk.Label(f3, text="Huishoudens:").grid(row=0, column=0, **pad)
        ttk.Label(f3, text="Opmerkingen:").grid(row=0, column=1, **pad)
        ttk.Label(f3, text="Loopafstand (m):").grid(row=0, column=2, **pad)

        self.e_huishoudens  = ttk.Entry(f3, width=10); self.e_huishoudens.grid(row=1, column=0, **pad)
        self.e_opmerkingen  = ttk.Entry(f3, width=28); self.e_opmerkingen.grid(row=1, column=1, **pad)
        self.e_loopafstand  = ttk.Entry(f3, width=10); self.e_loopafstand.grid(row=1, column=2, **pad)

        # --- rij 4: containers ---
        f4 = ttk.LabelFrame(self, text="Containers")
        f4.grid(row=3, column=0, columnspan=2, sticky="ew", **pad)

        fracties = ["Rest", "GFT", "PMD", "Papier", "Glas", "Textiel"]
        ttk.Label(f4, text="").grid(row=0, column=0)
        for i, fr in enumerate(fracties):
            ttk.Label(f4, text=fr, width=8, anchor="center").grid(row=0, column=i+1, **pad)

        ttk.Label(f4, text="Bestaand", width=10).grid(row=1, column=0, **pad)
        ttk.Label(f4, text="Nieuw",    width=10).grid(row=2, column=0, **pad)

        self.spinboxes = {}
        for j, fr in enumerate(fracties):
            sb_b = ttk.Spinbox(f4, from_=0, to=99, width=5)
            sb_b.set(0)
            sb_b.grid(row=1, column=j+1, **pad)
            sb_n = ttk.Spinbox(f4, from_=0, to=99, width=5)
            sb_n.set(0)
            sb_n.grid(row=2, column=j+1, **pad)
            self.spinboxes[fr] = (sb_b, sb_n)

        # --- rij 5: knoppen ---
        f5 = ttk.Frame(self)
        f5.grid(row=4, column=0, columnspan=2, pady=8)

        self.chk_screenshots_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(f5, text="Screenshots maken", variable=self.chk_screenshots_var).grid(row=0, column=0, padx=8)

        ttk.Button(f5, text="🌐  Open Websites",   command=self._open_browser,  width=20).grid(row=0, column=1, padx=8)
        ttk.Button(f5, text="🗑  Wissen",           command=self._clear,         width=14).grid(row=0, column=2, padx=8)
        ttk.Button(f5, text="📄  Maak Excel",       command=self._create_excel,  width=18).grid(row=0, column=3, padx=8)
        ttk.Button(f5, text="📂  Open Excel",       command=self._open_excel,    width=14).grid(row=0, column=4, padx=8)

        # --- statusbalk ---
        self.status_var = tk.StringVar(value="Klaar.")
        ttk.Label(self, textvariable=self.status_var, foreground="gray").grid(
            row=5, column=0, columnspan=2, sticky="w", padx=8, pady=2)

    # ---------------------------------------------------- GPS ophalen ---
    def _auto_coordinaten(self):
        """Haal coördinaten automatisch op als postcode en huisnummer zijn ingevuld."""
        if self.e_postcode.get() and self.e_huisnummer.get() and not self.e_coordinaten.get():
            threading.Thread(target=self._fetch_coordinaten, daemon=True).start()

    def _haal_coordinaten(self):
        """Handmatig coördinaten ophalen via de 📍 knop."""
        self.e_coordinaten.delete(0, tk.END)
        threading.Thread(target=self._fetch_coordinaten, daemon=True).start()

    def _fetch_coordinaten(self):
        try:
            import requests
            straat    = self.e_straat.get().strip()
            huisnr    = self.e_huisnummer.get().strip()
            toev      = self.e_toevoeging.get().strip()
            postcode  = self.e_postcode.get().strip().replace(" ", "")
            plaats    = self.cb_plaats.get().strip()

            query = f"{straat} {huisnr}{toev}, {postcode} {plaats}, Nederland"
            self._status("Coördinaten ophalen...")

            resp = requests.get(
                "https://nominatim.openstreetmap.org/search",
                params={"q": query, "format": "json", "limit": 1},
                headers={"User-Agent": "ProductbladGenerator/1.0"},
                timeout=10
            )
            results = resp.json()
            if results:
                lat = results[0]["lat"]
                lon = results[0]["lon"]
                coords = f"{lat}, {lon}"
                self.e_coordinaten.delete(0, tk.END)
                self.e_coordinaten.insert(0, coords)
                self._status(f"Coördinaten gevonden: {coords}")
            else:
                self._status("Geen coördinaten gevonden — vul handmatig in.")
        except Exception as e:
            self._status(f"Coördinaten ophalen mislukt: {e}")

    # --------------------------------------------------------- dropdowns ---
    def _fill_opdrachtgevers(self):
        namen = [o["naam"] for o in self.data["opdrachtgevers"]]
        self.cb_opdrachtgever["values"] = namen
        if namen:
            self.cb_opdrachtgever.current(0)
            self._update_plaats()

    def _update_plaats(self):
        idx = self.cb_opdrachtgever.current()
        plaatsen = [p["naam"] for p in self.data["opdrachtgevers"][idx]["plaatsen"]]
        self.cb_plaats["values"] = plaatsen
        if plaatsen:
            self.cb_plaats.current(0)
            self._update_wijk()

    def _update_wijk(self):
        idx_o = self.cb_opdrachtgever.current()
        idx_p = self.cb_plaats.current()
        wijken = self.data["opdrachtgevers"][idx_o]["plaatsen"][idx_p]["wijken"]
        self.cb_wijk["values"] = wijken
        if wijken:
            self.cb_wijk.current(0)

    # ------------------------------------------------------------ acties ---
    def _clear(self):
        for e in [self.e_straat, self.e_huisnummer, self.e_toevoeging,
                  self.e_postcode, self.e_coordinaten, self.e_huishoudens,
                  self.e_opmerkingen, self.e_loopafstand]:
            e.delete(0, tk.END)
        for sb_b, sb_n in self.spinboxes.values():
            sb_b.set(0); sb_n.set(0)

    def _status(self, txt):
        self.status_var.set(txt)
        self.update_idletasks()

    def _open_browser(self):
        self._status("Browser openen...")
        threading.Thread(target=self._open_browser_thread, daemon=True).start()

    def _open_browser_thread(self):
        try:
            from Dependencies import WebDriver as web
            chrome_driver = os.path.join(DEPS, "chromedriver.exe")
            straat     = self.e_straat.get()
            huisnr     = self.e_huisnummer.get()
            toev       = self.e_toevoeging.get()
            plaats     = self.cb_plaats.get()
            coords     = self.e_coordinaten.get()
            target_address = f"{straat} {huisnr}{toev}, {plaats}"

            self.web_driver = web.WebDriver(
                chrome_driver,
                f"https://bagviewer.kadaster.nl/lvbag/bag-viewer/?searchQuery={target_address}"
            )
            time.sleep(4)
            self.web_driver.new_tab("https://app.pdok.nl/viewer/", extra_wait=2.0)
            self.web_driver.new_search("ggcSearchInputId", target_address)
            self.google_url = f"https://www.google.nl/maps/place/{target_address}/data=!3m1!1e3"
            self.web_driver.new_tab(self.google_url, extra_wait=2.0)
            self.web_driver.new_tab("https://www.gpscoordinaten.nl/converteer-gps-coordinaten.php", extra_wait=2.5)
            self.web_driver.new_search("a-latlong", coords)
            joined = coords.replace(" ", "")
            self.web_driver.new_tab(f"https://www.google.com/maps/@?api=1&map_action=pano&viewpoint={joined}", extra_wait=2.0)
            self.web_driver.new_tab("https://afstandmeten.nl/", extra_wait=3.0)
            self.web_driver.new_search("qId", target_address)
            self._status("Websites geopend. Pas screenshots aan en klik op 'Maak Excel'.")
        except Exception as e:
            messagebox.showerror("Fout", f"Browser kon niet openen:\n{e}")
            self._status("Fout bij openen browser.")

    def _create_excel(self):
        self._status("Excel aanmaken...")
        threading.Thread(target=self._create_excel_thread, daemon=True).start()

    def _create_excel_thread(self):
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl import load_workbook
            from pyproj import Transformer

            # Locatie code (berekenen vóór screenshots, zodat we de map al klaar hebben)
            postcode = self.e_postcode.get().strip()
            letters = re.findall(r'\D+', postcode)
            pc_letters = letters[0].strip() if letters else postcode
            straat = self.e_straat.get().strip()
            huisnr = self.e_huisnummer.get().strip()
            self.location_code = f"{pc_letters}{huisnr} - {straat}"

            veilige_code     = re.sub(r'[\\/*?:"<>|]', '_', self.location_code).strip()
            veilige_gemeente = re.sub(r'[\\/*?:"<>|]', '_', self.cb_plaats.get()).strip()
            self.output_dir  = os.path.join(FILE_LOCATION, "Pythonwerk", veilige_gemeente, veilige_code)
            os.makedirs(self.output_dir, exist_ok=True)

            # Screenshots
            if self.chk_screenshots_var.get() and self.web_driver:
                try:
                    import pyscreeze
                    screenshot_names  = ['Kaart', 'Kadaster', 'Luchtfoto', 'GPS', 'Locatiefoto', 'Loopafstand']
                    image_regions     = [
                        [400, 350, 1095, 565],
                        [400, 350, 1300, 655],
                        [910, 210,  600, 790],
                        [175, 500, 1000, 387],
                        [300, 150, 1220, 720],
                        [270, 150, 1200, 700],
                    ]
                    handles = self.web_driver.web_driver.window_handles[:6]
                    for index, handle in enumerate(handles):
                        self.web_driver.web_driver.switch_to.window(handle)
                        time.sleep(2.0)
                        pyscreeze.screenshot(
                            os.path.join(self.output_dir, f"{screenshot_names[index]}.png"),
                            image_regions[index]
                        )
                except Exception:
                    self._status("Browser niet beschikbaar — screenshots overgeslagen.")

            # Workbook laden
            wb_path = os.path.join(DEPS, "PythonWerkProjectblad.xlsm")
            workbook = load_workbook(filename=wb_path, keep_vba=True)
            sheets = workbook.sheetnames

            info_sheet    = workbook[sheets[0]]
            general_sheet = workbook[sheets[1]]

            set_cell(info_sheet, 'D2', f"Gemeente: {self.cb_opdrachtgever.get()}")

            # Container nummers invullen
            fracties = ["Rest", "GFT", "PMD", "Papier", "Glas", "Textiel"]
            put_waarden = ["PUT", "PUT", "PUT", "Put"]

            bestaand_rij = 3
            teller_b = 1
            for fractie in fracties:
                sb_b, _ = self.spinboxes[fractie]
                for _ in range(int(sb_b.get())):
                    if bestaand_rij <= 7:
                        set_cell(info_sheet, f'A{bestaand_rij}', f"{teller_b} OOC {fractie}")
                        set_cell(info_sheet, f'B{bestaand_rij}', "OOC........")
                        set_cell(info_sheet, f'C{bestaand_rij}', "Bouwjaar")
                        set_cell(info_sheet, f'F{bestaand_rij}', "monolitisch")
                        for i, put in enumerate(put_waarden):
                            set_cell(info_sheet, f'E{bestaand_rij + i}', put)
                        bestaand_rij += 1
                        teller_b += 1

            nieuw_rij = 9
            teller_n = 1
            for fractie in fracties:
                _, sb_n = self.spinboxes[fractie]
                for _ in range(int(sb_n.get())):
                    if nieuw_rij <= 13:
                        set_cell(info_sheet, f'A{nieuw_rij}', f"{teller_n} OOC {fractie}")
                        set_cell(info_sheet, f'B{nieuw_rij}', "OOC........")
                        set_cell(info_sheet, f'C{nieuw_rij}', "Bouwjaar")
                        set_cell(info_sheet, f'F{nieuw_rij}', "monolitisch")
                        for i, put in enumerate(put_waarden):
                            set_cell(info_sheet, f'E{nieuw_rij + i}', put)
                        nieuw_rij += 1
                        teller_n += 1

            # Loopafstand
            loopafstand = self.e_loopafstand.get().strip().replace(",", ".")
            if loopafstand:
                try:
                    ld = float(loopafstand)
                    set_cell(general_sheet, 'C15', '<50 m' if ld < 50 else f"{int(ld)} m")
                except ValueError:
                    pass

            set_cell(general_sheet, 'G15', f"{self.e_huishoudens.get()} hh")

            # Coordinaten — robuuste parser (werkt met elk formaat)
            coords_raw = self.e_coordinaten.get().strip()
            if coords_raw:
                try:
                    # Haal alle getallen (met decimalen) uit de string
                    nums = re.findall(r'-?\d+\.\d+', coords_raw)
                    if len(nums) >= 2:
                        a, b = float(nums[0]), float(nums[1])
                        # Nederland: lat 50-54, lon 3-8
                        if 50 <= a <= 54 and 3 <= b <= 8:
                            lat, lon = a, b
                        elif 50 <= b <= 54 and 3 <= a <= 8:
                            lat, lon = b, a
                        else:
                            lat, lon = a, b
                        set_cell(general_sheet, 'C17', str(lat)[:10])
                        set_cell(general_sheet, 'E17', str(lon)[:10])
                        transformer = Transformer.from_crs("epsg:4326", "epsg:28992", always_xy=True)
                        x_rd, y_rd = transformer.transform(lon, lat)
                        set_cell(general_sheet, 'C16', round(x_rd))
                        set_cell(general_sheet, 'E16', round(y_rd))
                except Exception:
                    pass

            # Adres
            set_cell(general_sheet, 'B19', self.e_straat.get())
            set_cell(general_sheet, 'B20', self.e_postcode.get())
            set_cell(general_sheet, 'B21', self.cb_wijk.get())
            set_cell(general_sheet, 'F19', f"{self.e_huisnummer.get()}{self.e_toevoeging.get()}")
            set_cell(general_sheet, 'F20', self.cb_plaats.get())
            if self.google_url:
                set_cell(general_sheet, 'F21', f'=HYPERLINK("{self.google_url}", "Google")')
            set_cell(general_sheet, 'B42', self.e_opmerkingen.get())
            set_cell(general_sheet, 'B31', 'Hoogbouw in orde')

            # Data validaties
            dv1 = DataValidation(type="list", formula1="=Data!$A$1:$A$4")
            general_sheet.add_data_validation(dv1)
            dv1.add("C23:C28"); dv1.add("F23:F25"); dv1.add("H29:H30")
            dv2 = DataValidation(type="list", formula1="=Data!$B$1:$B$4")
            general_sheet.add_data_validation(dv2)
            dv2.add("B33")

            # Afbeeldingen
            screenshot_to_range = {
                'Locatiefoto': ('A68',  'H92'),
                'Kaart':       ('A94',  'H118'),
                'GPS':         ('A120', 'F130'),
                'Loopafstand': ('A201', 'H228'),
                'Kadaster':    ('A230', 'H255'),
                'Luchtfoto':   ('A267', 'H326'),
            }
            for name, (fc, tc) in screenshot_to_range.items():
                img_path = os.path.join(self.output_dir, f"{name}.png")
                if os.path.exists(img_path):
                    add_image_to_range(general_sheet, img_path, fc, tc)

            vaste = set(screenshot_to_range.keys()) | {'Logo', 'TransLogo'}
            extra = sorted([
                f for f in os.listdir(self.output_dir)
                if f.lower().endswith('.png') and os.path.splitext(f)[0] not in vaste
            ])
            if len(extra) >= 1:
                add_image_to_range(general_sheet, os.path.join(self.output_dir, extra[0]), 'A53', 'D65')
            if len(extra) >= 2:
                add_image_to_range(general_sheet, os.path.join(self.output_dir, extra[1]), 'E53', 'H65')

            # KLIC afbeelding zoeken op locatiecode
            klic_code = veilige_code.split(' - ')[0].strip()
            klic_bestand = None
            if os.path.isdir(KLIC):
                for f in os.listdir(KLIC):
                    name_lower = os.path.splitext(f)[0].lower()
                    if name_lower.startswith(klic_code.lower()) and f.lower().endswith(('.png', '.jpg', '.jpeg')):
                        klic_bestand = os.path.join(KLIC, f)
                        break
            if klic_bestand:
                add_image_to_range(general_sheet, klic_bestand, 'A135', 'H184')

            # Opslaan in output map (als .xlsm om VBA te behouden)
            out_xlsx = os.path.join(self.output_dir, f"{veilige_code}.xlsm")
            workbook.save(out_xlsx)

            self._status(f"Klaar! Bestand: {veilige_code}.xlsm")
            messagebox.showinfo("Klaar", f"Excel aangemaakt:\n{out_xlsx}")

        except Exception as e:
            import traceback
            messagebox.showerror("Fout", f"Excel kon niet worden aangemaakt:\n{e}\n\n{traceback.format_exc()}")
            self._status("Fout bij aanmaken Excel.")

    def _open_excel(self):
        if not self.location_code or not hasattr(self, 'output_dir'):
            messagebox.showwarning("Let op", "Eerst Excel aanmaken.")
            return
        veilige_code = re.sub(r'[\\/*?:"<>|]', '_', self.location_code).strip()
        path = os.path.join(self.output_dir, f"{veilige_code}.xlsm")
        if os.path.exists(path):
            os.startfile(path)
        else:
            messagebox.showerror("Niet gevonden", f"Bestand niet gevonden:\n{path}")


if __name__ == "__main__":
    app = App()
    app.mainloop()
