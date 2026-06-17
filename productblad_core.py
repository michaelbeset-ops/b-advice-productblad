# -*- coding: utf-8 -*-
"""
Headless kern-functie voor het genereren van een productblad-Excel.
Wordt aangeroepen door watcher.py (hot-folder).

Container JSON-formaat (nieuw, per container gespecificeerd):

    "containers": {
        "bestaand": [
            {
                "fractie":       "REST",
                "containernummer": "OOC12345",
                "type_put":      "TTC 5m3",
                "constructie":   "Monolitisch",
                "bouwjaar":      "2015",
                "vloer":         "Klapvloer"
            }
        ],
        "nieuw": [
            {
                "fractie":       "PMD",
                "containernummer": "",
                "type_put":      "TTC 5m3",
                "constructie":   "Monolitisch",
                "bouwjaar":      "",
                "vloer":         "Klapvloer"
            }
        ]
    }

Kolommen in de Excel per containerrij:
    A  — nummer + fractie       bijv. "1. OOC REST"
    B  — containernummer        bijv. "OOC12345"  of "OOC........."
    C  — status put             bijv. "Bestaand"
    D  — type betonput          bijv. "TTC 5m3"
    E  — "PUT"
    F  — constructie            bijv. "Monolitisch"
    G  — bouwjaar               bijv. "2015"
    H  — vloer                  bijv. "Klapvloer"
"""

import os
import re

FILE_LOCATION = os.path.dirname(os.path.realpath(__file__))
DEPS = os.path.join(FILE_LOCATION, "Dependencies")
KLIC = os.path.join(FILE_LOCATION, "KLIC")

# ---------------------------------------------------------------------------
# Presets: standaardwaarden per containertype
#
# Gebruik in JSON:  "preset": "OOC"  of  "preset": "BOC_PUT"  of  "preset": "BOC"
# Je kunt altijd een veld handmatig overschrijven door het mee te geven.
#
#  OOC      → ondergrondse container met betonput en klapvloer
#  BOC_PUT  → bovengrondse container MET betonput (GFT, PMD)
#  BOC      → bovengrondse container ZONDER betonput (OPK, Glas, Textiel, ...)
# ---------------------------------------------------------------------------
PRESETS = {
    "OOC": {
        "containernummer": "OOC",
        "type_put":        "TTC 5m3",
        "put":             "Put",
        "constructie":     "Monolitisch",
        "bouwjaar":        "Bouwjaar",
        "vloer":           "Klapvloer",
    },
    "BOC_PUT": {
        "containernummer": "BOC",
        "type_put":        "TTC 5m3",
        "put":             "Put",
        "constructie":     "Monolitisch",
        "bouwjaar":        "Bouwjaar",
        "vloer":           "",
    },
    "BOC": {
        "containernummer": "BOC",
        "type_put":        "",
        "put":             "",
        "constructie":     "Monolitisch",
        "bouwjaar":        "Bouwjaar",
        "vloer":           "",
    },
}


def set_cell(sheet, cell_ref, value):
    from openpyxl.utils import column_index_from_string
    col_str = re.match(r'([A-Z]+)', cell_ref.upper()).group(1)
    row = int(re.match(r'[A-Z]+(\d+)', cell_ref.upper()).group(1))
    col = column_index_from_string(col_str)
    for merged in sheet.merged_cells.ranges:
        if (merged.min_row <= row <= merged.max_row and
                merged.min_col <= col <= merged.max_col):
            sheet.cell(row=merged.min_row, column=merged.min_col).value = value
            return
    sheet[cell_ref] = value


def add_image_to_range(sheet, img_path, from_cell, to_cell):
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    from openpyxl.utils import column_index_from_string
    from openpyxl.drawing.image import Image

    def parse(ref):
        m = re.match(r'([A-Z]+)(\d+)', ref.upper())
        return column_index_from_string(m.group(1)) - 1, int(m.group(2)) - 1

    fc, fr = parse(from_cell)
    tc, tr = parse(to_cell)
    anchor = TwoCellAnchor()
    anchor._from = AnchorMarker(col=fc, colOff=0, row=fr, rowOff=0)
    anchor.to    = AnchorMarker(col=tc, colOff=0, row=tr, rowOff=0)
    anchor.editAs = 'oneCell'
    img = Image(img_path)
    img.anchor = anchor
    sheet.add_image(img)


def _vul_container_rij(sheet, rij: int, teller: int, container: dict, status_label: str):
    """
    Vul één containerrij in op het info-blad.
    Bestaand: rijen 3-7  |  Nieuw: rijen 9-13

    Werkt op twee manieren:
      1. Met preset:  {"preset": "OOC", "fractie": "REST"}
         → alle waarden komen uit PRESETS, je kunt velden nog losse overschrijven
      2. Volledig handmatig: alle velden zelf opgeven
    """
    # Preset laden als opgegeven
    preset_naam = str(container.get("preset", "")).upper()
    basis = dict(PRESETS.get(preset_naam, PRESETS["OOC"]))  # OOC als fallback

    # Handmatige velden overschrijven de preset
    for veld in ("containernummer", "type_put", "put", "constructie", "bouwjaar", "vloer"):
        if container.get(veld, "") not in ("", None):
            basis[veld] = str(container[veld])

    fractie     = str(container.get("fractie", "")).upper()
    container_prefix = preset_naam if preset_naam in PRESETS else "OOC"

    set_cell(sheet, f'A{rij}', f"{teller}. {container_prefix} {fractie}")

    # Kolom B heeft een Excel-formule die automatisch "OOC........." invult.
    # Alleen overschrijven als er een specifiek containernummer is opgegeven.
    specifiek_nr = str(container.get("containernummer", "")).strip()
    if specifiek_nr and specifiek_nr not in ("OOC", "BOC", "OOC.........", ""):
        set_cell(sheet, f'B{rij}', specifiek_nr)

    set_cell(sheet, f'C{rij}', status_label)
    set_cell(sheet, f'D{rij}', basis["type_put"])
    set_cell(sheet, f'E{rij}', basis["put"])
    set_cell(sheet, f'F{rij}', basis["constructie"])
    set_cell(sheet, f'G{rij}', basis["bouwjaar"])
    set_cell(sheet, f'H{rij}', basis["vloer"])


def generate_excel(data: dict) -> str:
    """
    Genereer een productblad-Excel op basis van data-dict uit input.json.
    Geeft het pad naar het aangemaakte .xlsm-bestand terug.
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl import load_workbook
    from pyproj import Transformer

    opdrachtgever = str(data.get("opdrachtgever", ""))
    plaats        = str(data.get("plaats", ""))
    wijk          = str(data.get("wijk", ""))
    straat        = str(data.get("straat", ""))
    huisnr        = str(data.get("huisnummer", ""))
    toev          = str(data.get("toevoeging", ""))
    postcode      = str(data.get("postcode", ""))
    coords_raw    = str(data.get("coordinaten", ""))
    huishoudens   = str(data.get("huishoudens", ""))
    opmerkingen   = str(data.get("opmerkingen", ""))
    loopafstand   = str(data.get("loopafstand", "")).replace(",", ".")
    containers    = data.get("containers", {})
    extra_fotos   = data.get("extra_fotos", [])

    bestaand_lijst = containers.get("bestaand", [])
    nieuw_lijst    = containers.get("nieuw",    [])

    # Locatiecode en output-map
    letters = re.findall(r'\D+', postcode)
    pc_letters = letters[0].strip() if letters else postcode
    location_code    = f"{pc_letters}{huisnr} - {straat}"
    veilige_code     = re.sub(r'[\\/*?:"<>|]', '_', location_code).strip()
    veilige_gemeente = re.sub(r'[\\/*?:"<>|]', '_', plaats).strip()
    output_dir = os.path.join(FILE_LOCATION, "Pythonwerk", veilige_gemeente, veilige_code)
    os.makedirs(output_dir, exist_ok=True)

    # Workbook laden
    wb_path  = os.path.join(DEPS, "PythonWerkProjectblad.xlsm")
    workbook = load_workbook(filename=wb_path, keep_vba=True)
    sheets   = workbook.sheetnames

    info_sheet    = workbook[sheets[0]]
    general_sheet = workbook[sheets[1]]

    # Gemeente op Projectomschrijving (sheet 0)
    set_cell(info_sheet, 'D2', f"Gemeente: {opdrachtgever}")

    # Containers op Projectblad (sheet 1) — rijen 3-7 bestaand, 9-13 nieuw
    for teller, container in enumerate(bestaand_lijst[:5], start=1):
        _vul_container_rij(general_sheet, rij=2 + teller, teller=teller,
                           container=container, status_label="Bestaand")

    for teller, container in enumerate(nieuw_lijst[:5], start=1):
        _vul_container_rij(general_sheet, rij=8 + teller, teller=teller,
                           container=container, status_label="Nieuw")

    # Loopafstand
    if loopafstand:
        try:
            ld = float(loopafstand)
            set_cell(general_sheet, 'C15', '<50 m' if ld < 50 else f"{int(ld)} m")
        except ValueError:
            pass

    set_cell(general_sheet, 'G15', f"{huishoudens} hh")

    # Coordinaten
    if coords_raw:
        try:
            nums = re.findall(r'-?\d+\.\d+', coords_raw)
            if len(nums) >= 2:
                a, b = float(nums[0]), float(nums[1])
                if 50 <= a <= 54 and 3 <= b <= 8:
                    lat, lon = a, b
                elif 50 <= b <= 54 and 3 <= a <= 8:
                    lat, lon = b, a
                else:
                    lat, lon = a, b
                set_cell(general_sheet, 'C17', str(lat)[:10])
                set_cell(general_sheet, 'E17', str(lon)[:10])
                transformer = Transformer.from_crs("epsg:4326", "epsg:28992", always_xy=True)
                x_rd, y_rd = transformer.transform(lon, lat)
                set_cell(general_sheet, 'C16', round(x_rd))
                set_cell(general_sheet, 'E16', round(y_rd))
        except Exception:
            pass

    # Adres
    google_url = (f"https://www.google.nl/maps/place/"
                  f"{straat} {huisnr}{toev}, {plaats}/data=!3m1!1e3")
    set_cell(general_sheet, 'B19', straat)
    set_cell(general_sheet, 'B20', postcode)
    set_cell(general_sheet, 'B21', wijk)
    set_cell(general_sheet, 'F19', f"{huisnr}{toev}")
    set_cell(general_sheet, 'F20', plaats)
    set_cell(general_sheet, 'F21', f'=HYPERLINK("{google_url}", "Google")')
    set_cell(general_sheet, 'B42', opmerkingen)
    set_cell(general_sheet, 'B31', 'Hoogbouw in orde')

    # Data validaties
    dv1 = DataValidation(type="list", formula1="=Data!$A$1:$A$4")
    general_sheet.add_data_validation(dv1)
    dv1.add("C23:C28"); dv1.add("F23:F25"); dv1.add("H29:H30")
    dv2 = DataValidation(type="list", formula1="=Data!$B$1:$B$4")
    general_sheet.add_data_validation(dv2)
    dv2.add("B33")

    # Screenshots
    screenshot_to_range = {
        'Locatiefoto': ('A68',  'H92'),
        'Kaart':       ('A94',  'H118'),
        'GPS':         ('A120', 'F130'),
        'Loopafstand': ('A201', 'H228'),
        'Kadaster':    ('A230', 'H255'),
        'Luchtfoto':   ('A267', 'H326'),
    }
    for name, (fc, tc) in screenshot_to_range.items():
        img_path = os.path.join(output_dir, f"{name}.png")
        if os.path.exists(img_path):
            add_image_to_range(general_sheet, img_path, fc, tc)

    # Extra foto's
    vaste = set(screenshot_to_range.keys()) | {'Logo', 'TransLogo'}
    if extra_fotos:
        candidates = [p for p in extra_fotos if os.path.exists(p)]
    else:
        candidates = sorted([
            os.path.join(output_dir, f)
            for f in os.listdir(output_dir)
            if f.lower().endswith('.png') and os.path.splitext(f)[0] not in vaste
        ])
    if len(candidates) >= 1:
        add_image_to_range(general_sheet, candidates[0], 'A53', 'D65')
    if len(candidates) >= 2:
        add_image_to_range(general_sheet, candidates[1], 'E53', 'H65')

    # KLIC afbeelding
    klic_code = veilige_code.split(' - ')[0].strip()
    if os.path.isdir(KLIC):
        for f in os.listdir(KLIC):
            if (os.path.splitext(f)[0].lower().startswith(klic_code.lower())
                    and f.lower().endswith(('.png', '.jpg', '.jpeg'))):
                add_image_to_range(general_sheet, os.path.join(KLIC, f), 'A135', 'H184')
                break

    out_path = os.path.join(output_dir, f"{veilige_code}.xlsm")
    workbook.save(out_path)
    return out_path
